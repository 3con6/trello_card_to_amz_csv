import asyncio
from b_trello import (
    get_trello_cards,
    get_attachments_in_card,
    move_cart_to_list,
    comment_to_cart,
    create_card_on_list,
    create_attachment_on_card,
    download_trello_attachments
)
from discord import upload_discord
from process_img import add_watermark
import sys
from datetime import datetime
import time
import os
import openpyxl



if sys.version_info[0] == 3 and sys.version_info[1] >= 8 and sys.platform.startswith('win'):
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())


async def add_watermark_from_trello_to_discord(url):
    trello_img = await download_trello_attachments(url)
    watermark_img = await add_watermark(trello_img)
    url = await upload_discord(watermark_img,'name')
    return url




async def get_description_and_attachemnts_from_card(card_id,trello_des,name):
    attachments = await get_attachments_in_card(card_id)
    all_imgs = []
    data = {'name':name,'id':card_id}
    for attachment in attachments:
        if '.txt'in attachment['name']:
            text = await download_trello_attachments(attachment['url'])
            data['description'] = text.decode("utf-8")
        elif '.jpg' in attachment['name'] or '.jpeg' in attachment['name'] or '.png' in attachment['name']: all_imgs.append(attachment)
    imgs = [attachment['url'] for attachment in all_imgs]
    trello_des_split = trello_des.splitlines()
    data['imgs'] = imgs
    for split in trello_des_split:
        if split != '':
            if '”' in split:
                split_ = split.split(': ”')
                print(split_)
                value = split_[1].replace('”','').replace('“','')    
            elif '“' in split:
                split_ = split.split(': “')
                value = split_[1].replace('“','')    
            key = split_[0].lower().replace(' ','_')
            data[key] = value
    return data      


async def do_fail_card(card_id,reason):
    await comment_to_cart(card_id,reason)
    await move_cart_to_list(card_id,'645b558b4fcd7e1c5eed6767')



async def generate_row_from_card(card, template_cards_dict, product_template):
    template_card_dict = template_cards_dict.get(product_template, None)
    if template_card_dict:
        row = [
            template_card_dict["product_type"],
            card["name"],
            template_card_dict["brand_name"],
            "",
            card["title"],
            template_card_dict["sub_category"],
            "",
            template_card_dict["description"],
            template_card_dict["price"],
            template_card_dict["quantity"],
        ]
        
        all_imgs = card["imgs"] + template_card_dict["imgs"]
        if len(all_imgs) > 9:
            await do_fail_card(
                card["id"],
                reason="Số lượng ảnh vượt quá 9 ảnh bao gồm cả ảnh trong template",
            )
            return None
        all_urls = await asyncio.gather(
            *(add_watermark_from_trello_to_discord(img) for img in all_imgs)
        )
        while len(all_urls) < 9:
            all_urls.append("")
        empty_space = ["" for _ in range(26)]
        new_empty_space1 = ["" for _ in range(37)]
        new_empty_space2 = ["" for _ in range(31)]
        row += (
            all_urls
            + empty_space
            + ["", "", card["tag"], "", template_card_dict["bullet_poll_1"], template_card_dict["bullet_poll_2"], template_card_dict["bullet_poll_3"], template_card_dict["bullet_poll_4"], template_card_dict["bullet_poll_5"]]
            + new_empty_space1
            + ["Not Applicable", "Unknown", "Unknown", "Unknown", "Unknown"]
            + new_empty_space2
            + [template_card_dict["handling_time"], "", "", "", "", "", "", "", "No"]
        )
        return row



async def process_ready_card(card,template_cards_dict):
    await move_cart_to_list(card['id'],'645b53367806a33f0d400ccf')
    split_name = card['name'].split('-')
    if len(split_name) != 5:
        await do_fail_card(card['id'],reason='Sai cách đặt SKU')
        return
    product_template = split_name[3]
    template_card_dict = template_cards_dict.get(product_template,None)
    if template_card_dict:
        row = await generate_row_from_card(card,template_cards_dict,product_template) 
        await move_cart_to_list(card['id'],'645b533c05ca5f5336ee3c3a')
        return row
    else:
        await do_fail_card(card['id'],reason='Mã sản phẩm SKU không có trong cột Template')
        return 


def convert_list_to_dict(template_cards_data):
    template_cards_dict = {}
    for item in template_cards_data:
        name = item['name']
        template_cards_dict[name] = item
    return template_cards_dict    



async def process_trello():
    ready_list_id = '645b54085f9172e970ca2037'
    ready_cards =  await get_trello_cards(ready_list_id, 'name,desc')
    if ready_cards != []:
        template_list_id =  '645b53277955359f2a5f7b2e'
        template_cards =  await get_trello_cards(template_list_id, 'name,desc')
        template_cards_data = await asyncio.gather(*(get_description_and_attachemnts_from_card(card['id'],card['desc'],card['name']) for card in template_cards))
        template_cards_dict = convert_list_to_dict(template_cards_data)
        ready_cards_data = await asyncio.gather(*(get_description_and_attachemnts_from_card(card['id'],card['desc'],card['name']) for card in ready_cards))
        rows_ready_card = await asyncio.gather(*(process_ready_card(card,template_cards_dict) for card in ready_cards_data))
        results = [row for row in rows_ready_card if row != None]              
        return results    


def clone_workbook_and_fill_data(workbook,data):
    new_wb = openpyxl.Workbook()
    for sheet in workbook.sheetnames:
        new_sheet = new_wb.create_sheet(sheet)
        for row in list(workbook[sheet].values):
            new_sheet.append(row)
        if sheet == 'Template':
            for row_ in data:
                new_sheet.append(row_)
    return new_wb            

def create_workbook_from_data(data):
    timestamp = time.time()
    today = datetime.today().strftime('%d-%m-%Y')
    name_xlsx = f'{today}_{timestamp}.xlsm'
    template_wb = openpyxl.load_workbook('template.xlsm',read_only=True)
    new_wb = clone_workbook_and_fill_data(template_wb,data)
    new_wb.save(name_xlsx)
    return name_xlsx

def get_bytes_from_file(name):
    with open(name, 'rb') as f:
        data_xlsx = f.read()
        return data_xlsx



async def main():
    data = await process_trello()
    if data and data != [] :
        name =  create_workbook_from_data(data)
        data_xlsx = get_bytes_from_file(name)
        card = await create_card_on_list('645b53a403be1bf0257efa2f',name)    
        await create_attachment_on_card(card['id'],name,data_xlsx)
        os.remove(name)

while True:
    asyncio.run(main())
    time.sleep(5)
